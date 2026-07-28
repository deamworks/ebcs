# ════════════════════════════════════════════════════════
# services/import_service.py
# ════════════════════════════════════════════════════════
#
# ไฟล์ Excel จริงที่รองรับ:
#
# [taxpayer] รายงานข้อมูลผู้ประกอบการ_*.xlsx
#   row1 = ว่าง
#   row2 = ชื่อรายงาน (merged)
#   row3 = HEADER → ลำดับ | เลขที่ผู้เสียภาษี | ผู้ประกอบการ
#                    | รอบระยะเวลาบัญชี | วันครบกำหนดชำระ | ...
#   row4+ = ข้อมูล
#
# [licensee] รายงานข้อมูลใบอนุญาต_*.xlsx
#   row1 = ว่าง
#   row2 = ชื่อรายงาน (merged)
#   row3 = HEADER → ลำดับ | ปี | รหัสอ้างอิง | ประเภท | รอบ | สถานะ
#                    | เลขประจำตัวผู้เสียภาษี | รายชื่อผู้รับใบอนุญาต
#                    | จำนวนใบอนุญาต | เลขที่ใบอนุญาต | ประเภทใบอนุญาต
#                    | วันที่เริ่มต้น | วันที่สิ้นสุด | สถานะใบอนุญาต
#                    | รายได้จากใบอนุญาต | ค่าลดหย่อน | เงินนำส่งเข้ากองทุน
#                    | ภาษีมูลค่าเพิ่ม | เงินเพิ่ม | จำนวนเงินรายปีนำส่งเข้ากองทุนสุทธิ
#   row4+ = ข้อมูล
# ════════════════════════════════════════════════════════

import io
import re
from datetime import datetime, date

import openpyxl

EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")


# ────────────────────────────────────────────────────────
# Helpers
# ────────────────────────────────────────────────────────

def _to_date(val):
    """
    แปลงค่าต่างๆ → date object (ค.ศ.)
    รองรับ: datetime, date, 'DD/MM/YYYY' (BE/CE), 'YYYY-MM-DD'
    """
    if val is None:
        return None
    if isinstance(val, datetime):
        d = val.date()
        return date(d.year - 543, d.month, d.day) if d.year > 2400 else d
    if isinstance(val, date):
        return date(val.year - 543, val.month, val.day) if val.year > 2400 else val
    s = str(val).strip()
    if not s:
        return None
    m = re.match(r"(\d{1,2})[/\-](\d{1,2})[/\-](\d{4})", s)
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if y > 2400:
            y -= 543
        try:
            return date(y, mo, d)
        except ValueError:
            return None
    try:
        return date.fromisoformat(s[:10])
    except ValueError:
        return None


def _parse_period(raw):
    """
    '01/01/2569 - 31/12/2569' หรือ '01/01/2569-31/12/2569'
    → (period_start: date, period_end: date)
    """
    if not raw:
        return None, None
    # แทนเส้นประหลายแบบ แล้ว split
    s = str(raw).strip().replace("–", "-").replace("—", "-")
    parts = re.split(r"\s*-\s*", s, maxsplit=1)
    if len(parts) == 2:
        return _to_date(parts[0].strip()), _to_date(parts[1].strip())
    return None, None


def _fiscal_year_be(period_end):
    """date ค.ศ. → ปี พ.ศ."""
    if period_end is None:
        return None
    return period_end.year + 543


def _num(val):
    try:
        return float(val) if val not in (None, "") else 0.0
    except (ValueError, TypeError):
        return 0.0


def _clean_tax_id(val):
    """แปลง float/string → string ตัวเลขล้วน"""
    s = str(val or "").strip().replace("-", "").replace(" ", "")
    if "." in s:
        s = s.split(".")[0]
    return s


def _read_data_rows(file_stream, header_row: int):
    """
    อ่าน xlsx คืน (headers: list[str], data: list[list])
    header_row = หมายเลขแถว Excel (1-based) ที่เป็น header
    data = แถวถัดไปทั้งหมดที่ไม่ว่างทั้งแถว
    """
    if isinstance(file_stream, (bytes, bytearray)):
        file_stream = io.BytesIO(file_stream)
    wb = openpyxl.load_workbook(file_stream, data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()
    headers = [str(c).strip() if c is not None else "" for c in all_rows[header_row - 1]]
    data = [
        list(r) for r in all_rows[header_row:]   # ข้ามแถว header
        if any(c is not None and str(c).strip() for c in r)
    ]
    return headers, data


# ════════════════════════════════════════════════════════
# TAXPAYER  →  taxpayer_master
# ════════════════════════════════════════════════════════
# col index (จาก header แถว 3):
#  0  ลำดับ
#  1  เลขที่ผู้เสียภาษี   → tax_id
#  2  ผู้ประกอบการ         → operator_name
#  3  รอบระยะเวลาบัญชี    → period_start, period_end, fiscal_year (BE)
#  4  วันครบกำหนดชำระ     → due_date
#  5  ที่อยู่ที่ติดต่อได้  → address (รวม col 5-9)
#  6  ตำบล
#  7  อำเภอ
#  8  จังหวัด
#  9  รหัสไปรษณีย์
# 10  คำนำหน้า            (ไม่ใช้ — ชื่อรวมอยู่ใน col 2 แล้ว)
# col 11+ = EMS/เตือน → ไม่ import

def _parse_taxpayer_row(row, excel_row_num):
    errs = []

    # tax_id
    tax_id = _clean_tax_id(row[1] if len(row) > 1 else None)
    if not tax_id or not tax_id.isdigit() or len(tax_id) != 13:
        errs.append(f"เลขที่ผู้เสียภาษี '{tax_id}' ต้องเป็นตัวเลข 13 หลัก")

    # operator_name
    operator_name = str(row[2]).strip() if len(row) > 2 and row[2] else ""
    if not operator_name:
        errs.append("ชื่อผู้ประกอบการว่าง")

    # period → period_start, period_end, fiscal_year
    period_raw = str(row[3]).strip() if len(row) > 3 and row[3] else ""
    period_start, period_end = _parse_period(period_raw)
    fiscal_year = _fiscal_year_be(period_end)
    if not fiscal_year:
        errs.append(f"รอบระยะเวลาบัญชี '{period_raw}' อ่านไม่ได้")

    # due_date
    due_date = _to_date(row[4]) if len(row) > 4 and row[4] else None

    # address (รวม col 5-9)
    addr_parts = [str(row[i]).strip() for i in range(5, 10) if len(row) > i and row[i]]
    address = " ".join(p for p in addr_parts if p)

    if errs:
        return None, {"row": excel_row_num, "message": ", ".join(errs)}

    return {
        "tax_id":        tax_id,
        "operator_name": operator_name,
        "period_start":  period_start,
        "period_end":    period_end,
        "fiscal_year":   fiscal_year,
        "due_date":      due_date,
        "address":       address,
    }, None


def parse_taxpayer_excel(file_stream):
    """
    อ่านไฟล์ Excel ผู้ประกอบการ (header แถว 3, ข้อมูลแถว 4+)
    คืน: (rows: list[dict], errors: list[dict])
    """
    try:
        _, data = _read_data_rows(file_stream, header_row=3)
    except Exception as e:
        return [], [{"row": 0, "message": f"เปิดไฟล์ไม่ได้: {e}"}]

    rows, errors = [], []
    for idx, row in enumerate(data, start=4):   # Excel row 4 เป็นต้นไป
        rec, err = _parse_taxpayer_row(row, idx)
        if err:
            errors.append(err)
        else:
            rows.append(rec)
    return rows, errors


def _get_next_ref_no(cur, fiscal_year_be: int) -> str:
    """
    Generate ref_no อัตโนมัติ รูปแบบ BK{ปี 2 หลัก}{เลขลำดับ 5 หลัก}
    เช่น BK6800001, BK6800002
    ใช้ ref_no_counters table (unique key = prefix) เพื่อป้องกันเลขซ้ำ
    แม้ import พร้อมกัน — ตาราง ref_no_counters ไม่มีคอลัมน์ fiscal_year
    จึงใช้ prefix ("BK68") เป็น key ของตัวนับแทน
    """
    year_2digit = str(fiscal_year_be)[-2:]   # 2568 → "68"
    prefix = f"BK{year_2digit}"

    # INSERT ... ON DUPLICATE KEY UPDATE เพื่อ upsert แถว counter
    cur.execute("""
        INSERT INTO ref_no_counters (prefix, last_seq)
        VALUES (%s, 1)
        ON DUPLICATE KEY UPDATE last_seq = last_seq + 1
    """, (prefix,))

    cur.execute(
        "SELECT last_seq FROM ref_no_counters WHERE prefix = %s",
        (prefix,)
    )
    row = cur.fetchone()
    seq = row["last_seq"] if isinstance(row, dict) else row[0]
    return f"{prefix}{seq:05d}"


def import_taxpayers(db, rows):
    """
    Upsert taxpayer_master  unique: (tax_id, fiscal_year)
    - INSERT: generate ref_no อัตโนมัติ (NBTC-{ปีพ.ศ.}-{เลข 4 หลัก})
    - UPDATE: ไม่แตะ ref_no ที่มีอยู่แล้ว
    คืน: {"inserted": n, "updated": n, "snapshot": [...]}
    snapshot เก็บค่าก่อนนำเข้าของแต่ละแถว (old=None ถ้าเป็นแถวใหม่) สำหรับ rollback
    """
    inserted = updated = 0
    snapshot = []
    cur = db.cursor()
    for row in rows:
        cur.execute(
            "SELECT * FROM taxpayer_master WHERE tax_id=%s AND fiscal_year=%s",
            (row["tax_id"], row["fiscal_year"])
        )
        existing = cur.fetchone()
        key = {"tax_id": row["tax_id"], "fiscal_year": row["fiscal_year"]}

        if existing:
            snapshot.append({"table": "taxpayer_master", "key": key, "old": existing})
            cur.execute("""
                UPDATE taxpayer_master SET
                    operator_name = %s,
                    period_start  = %s,
                    period_end    = %s,
                    due_date      = %s,
                    address       = %s,
                    updated_at    = NOW()
                WHERE tax_id=%s AND fiscal_year=%s
            """, (
                row["operator_name"], row["period_start"],
                row["period_end"],   row["due_date"],
                row["address"],
                row["tax_id"],       row["fiscal_year"],
            ))
            updated += 1
        else:
            snapshot.append({"table": "taxpayer_master", "key": key, "old": None})
            # fiscal_year ใน row เป็น CE → แปลงเป็น BE สำหรับ ref_no
            fiscal_year_ce = row["fiscal_year"]
            fiscal_year_be = fiscal_year_ce + 543 if fiscal_year_ce < 2400 else fiscal_year_ce
            ref_no = _get_next_ref_no(cur, fiscal_year_be)

            cur.execute("""
                INSERT INTO taxpayer_master
                    (tax_id, operator_name, period_start, period_end,
                     fiscal_year, due_date, address, ref_no)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
            """, (
                row["tax_id"],       row["operator_name"],
                row["period_start"], row["period_end"],
                row["fiscal_year"],  row["due_date"],
                row["address"],      ref_no,
            ))
            inserted += 1
    db.commit()
    cur.close()
    return {"inserted": inserted, "updated": updated, "snapshot": snapshot}


# ════════════════════════════════════════════════════════
# LICENSEE  →  licenses
# ════════════════════════════════════════════════════════
# col index (จาก header แถว 3):
#  0  ลำดับ
#  1  ปี (BE)                    → fiscal_year (เก็บเป็น พ.ศ. ตรงกับ taxpayer_master)
#  2  รหัสอ้างอิง                → ref_code
#  3  ประเภท                     → sub_type
#  4  รอบ                        → round_type
#  5  สถานะ                      → sub_status
#  6  เลขประจำตัวผู้เสียภาษี     → tax_id
#  7  รายชื่อผู้รับใบอนุญาต      → company_name
#  8  จำนวนใบอนุญาต              → license_count
#  9  เลขที่ใบอนุญาต             → license_no  (unique key)
# 10  ประเภทใบอนุญาต             → license_type
# 11  วันที่เริ่มต้น              → license_start
# 12  วันที่สิ้นสุด               → license_end
# 13  สถานะใบอนุญาต              → license_status
# 14  รายได้จากใบอนุญาต          → income
# 15  ค่าลดหย่อน                 → deduction
# 16  เงินนำส่งเข้ากองทุน        → fund_amount
# 17  ภาษีมูลค่าเพิ่ม            → vat
# 18  เงินเพิ่ม                  → surcharge
# 19  จำนวนเงินรายปีนำส่งสุทธิ  → net_amount

# ไฟล์ Excel มักมี license_status เป็นภาษาไทย ต้อง map เป็น ENUM ก่อน insert ไม่งั้น MySQL error
_LICENSE_STATUS_MAP = {
    "active": "active", "ended": "ended", "cancelled": "cancelled", "revoked": "revoked",
    "ปกติ": "active", "ได้รับอนุญาต": "active",
    "สิ้นสุด": "ended", "สิ้นสุดระยะเวลา": "ended", "สิ้นสุดระยะเวลาอนุญาต": "ended",
    "ยกเลิก": "cancelled", "ยกเลิกประกอบกิจการ": "cancelled",
    "เพิกถอน": "revoked", "เพิกถอนใบอนุญาต": "revoked",
}


def _normalize_license_status(raw):
    """
    แปลงค่าคอลัมน์ "สถานะใบอนุญาต" (อาจเป็นภาษาไทยหรือค่า ENUM ตรงๆ)
    → ค่า ENUM จริง หรือคืน error message ถ้าไม่รู้จัก
    ค่าว่าง → ใช้ 'active' (ตรงกับ DEFAULT ของคอลัมน์นี้ใน schema)
    """
    if raw is None or str(raw).strip() == "":
        return "active", None
    key = str(raw).strip()
    mapped = _LICENSE_STATUS_MAP.get(key) or _LICENSE_STATUS_MAP.get(key.lower())
    if mapped:
        return mapped, None
    return None, (
        f"สถานะใบอนุญาต '{raw}' ไม่รู้จัก "
        "(ต้องเป็นหนึ่งใน: ปกติ, สิ้นสุด, ยกเลิก, เพิกถอน)"
    )


def _parse_licensee_row(row, excel_row_num):
    errs = []

    def _s(i): return str(row[i]).strip() if len(row) > i and row[i] is not None else None

    # tax_id
    tax_id = _clean_tax_id(row[6] if len(row) > 6 else None)
    if not tax_id or not tax_id.isdigit() or len(tax_id) != 13:
        errs.append(f"เลขที่ผู้เสียภาษี '{tax_id}' ต้องเป็นตัวเลข 13 หลัก")

    # license_no
    license_no = _s(9) or ""
    if not license_no:
        errs.append("เลขที่ใบอนุญาตว่าง")

    # [FIX] fiscal_year เก็บเป็น พ.ศ. ให้ตรงกับ taxpayer_master.fiscal_year เสมอ
    try:
        fiscal_year_be = int(str(row[1]).split(".")[0].strip()) if len(row) > 1 and row[1] else None
    except (TypeError, ValueError):
        fiscal_year_be = None
        errs.append(f"ปี '{row[1] if len(row) > 1 else ''}' อ่านไม่ได้")

    # license_status: ต้อง map ภาษาไทย → ENUM literal ก่อน ไม่งั้น insert ล้มเหลว
    license_status, status_err = _normalize_license_status(_s(13))
    if status_err:
        errs.append(status_err)

    if errs:
        return None, {"row": excel_row_num, "message": ", ".join(errs)}

    return {
        "tax_id":         tax_id,
        "fiscal_year":    fiscal_year_be,
        "ref_code":       _s(2),
        "sub_type":       _s(3),
        "round_type":     _s(4),
        "sub_status":     _s(5),
        "company_name":   _s(7),
        "license_count":  int(row[8]) if len(row) > 8 and row[8] else None,
        "license_no":     license_no,
        "license_type":   _s(10),
        "license_start":  _to_date(row[11]) if len(row) > 11 else None,
        "license_end":    _to_date(row[12]) if len(row) > 12 else None,
        "license_status": license_status,
        "income":         _num(row[14]) if len(row) > 14 else 0.0,
        "deduction":      _num(row[15]) if len(row) > 15 else 0.0,
        "fund_amount":    _num(row[16]) if len(row) > 16 else 0.0,
        "vat":            _num(row[17]) if len(row) > 17 else 0.0,
        "surcharge":      _num(row[18]) if len(row) > 18 else 0.0,
        "net_amount":     _num(row[19]) if len(row) > 19 else 0.0,
    }, None


def parse_licensee_excel(file_stream):
    """
    อ่านไฟล์ Excel ใบอนุญาต (header แถว 3, ข้อมูลแถว 4+)
    คืน: (rows: list[dict], errors: list[dict])
    """
    try:
        _, data = _read_data_rows(file_stream, header_row=3)
    except Exception as e:
        return [], [{"row": 0, "message": f"เปิดไฟล์ไม่ได้: {e}"}]

    rows, errors = [], []
    for idx, row in enumerate(data, start=4):
        rec, err = _parse_licensee_row(row, idx)
        if err:
            errors.append(err)
        else:
            rows.append(rec)
    return rows, errors


def import_licensees(db, rows):
    """
    Upsert licensee_master  unique: license_no
    คืน: {"inserted": n, "updated": n, "snapshot": [...]}
    snapshot เก็บค่าก่อนนำเข้าของแต่ละแถว (old=None ถ้าเป็นแถวใหม่) สำหรับ rollback
    """
    inserted = updated = 0
    snapshot = []
    cur = db.cursor()
    for row in rows:
        cur.execute("SELECT * FROM licensee_master WHERE license_no=%s", (row["license_no"],))
        existing = cur.fetchone()
        key = {"license_no": row["license_no"]}
        if existing:
            snapshot.append({"table": "licensee_master", "key": key, "old": existing})
            cur.execute("""
                UPDATE licensee_master SET
                    tax_id=%s, fiscal_year=%s, ref_code=%s,
                    sub_type=%s, round_type=%s, sub_status=%s,
                    company_name=%s, license_count=%s,
                    licensee_type=%s, start_date=%s, end_date=%s,
                    license_status=%s,
                    income=%s, deduction=%s, fund_amount=%s,
                    vat=%s, surcharge=%s, net_amount=%s,
                    updated_at=NOW()
                WHERE license_no=%s
            """, (
                row["tax_id"],        row["fiscal_year"],  row["ref_code"],
                row["sub_type"],      row["round_type"],   row["sub_status"],
                row["company_name"],  row["license_count"],
                row["license_type"],  row["license_start"],row["license_end"],
                row["license_status"],
                row["income"],        row["deduction"],    row["fund_amount"],
                row["vat"],           row["surcharge"],    row["net_amount"],
                row["license_no"],
            ))
            updated += 1
        else:
            snapshot.append({"table": "licensee_master", "key": key, "old": None})
            cur.execute("""
                INSERT INTO licensee_master (
                    tax_id, fiscal_year, ref_code, sub_type, round_type,
                    sub_status, company_name, license_count, license_no,
                    licensee_type, start_date, end_date, license_status,
                    income, deduction, fund_amount, vat, surcharge, net_amount
                ) VALUES (
                    %s,%s,%s,%s,%s,
                    %s,%s,%s,%s,
                    %s,%s,%s,%s,
                    %s,%s,%s,%s,%s,%s
                )
            """, (
                row["tax_id"],        row["fiscal_year"],  row["ref_code"],
                row["sub_type"],      row["round_type"],   row["sub_status"],
                row["company_name"],  row["license_count"],row["license_no"],
                row["license_type"],  row["license_start"],row["license_end"],
                row["license_status"],
                row["income"],        row["deduction"],    row["fund_amount"],
                row["vat"],           row["surcharge"],    row["net_amount"],
            ))
            inserted += 1
    db.commit()
    cur.close()
    return {"inserted": inserted, "updated": updated, "snapshot": snapshot}


# ════════════════════════════════════════════════════════
# CONTACTS  →  อัปเดต address + email ใน taxpayer_master
# ════════════════════════════════════════════════════════
# โครงสร้างไฟล์จริง (header แถว 5, ข้อมูลเริ่มแถว 6) 32 คอลัมน์:
#  col 0   ลำดับ
#  col 1   เลขประจำตัวผู้เสียภาษี  → tax_id
#  col 2   รอบบัญชี
#  col 3   วันที่เริ่มต้น
#  col 4   วันที่สิ้นสุด
#  col 5   รายชื่อผู้รับใบอนุญาต   → operator_name (อ้างอิง)
#  col 6   (ว่าง)
#  col 7   สถานะ
#  col 8   ประเภทและโครงสร้าง
#  col 9   ประเภทและโครงสร้าง
#  col 10  หมายเหตุ
#  col 11  ที่อยู่ตามระบบใบอนุญาต
#  col 12  ตำบล (ใบอนุญาต)
#  col 13  อำเภอ (ใบอนุญาต)
#  col 14  จังหวัด (ใบอนุญาต)
#  col 15  รหัสไปรษณีย์ (ใบอนุญาต)
#  col 16  ที่อยู่ที่แสดงในใบเสร็จ
#  col 17  ที่อยู่ออกใบเสร็จ        → address_receipt
#  col 18  ตำบล (ใบเสร็จ)
#  col 19  อำเภอ (ใบเสร็จ)
#  col 20  จังหวัด (ใบเสร็จ)
#  col 21  รหัสไปรษณีย์ (ใบเสร็จ)
#  col 22  โทรศัพท์ (ใบเสร็จ)
#  col 23  อีเมล (ใบเสร็จ)
#  col 24  ที่อยู่ที่ติดต่อได้(ชส.)  → address  ← ใช้อันนี้
#  col 25  ตำบล (ชส.)
#  col 26  อำเภอ (ชส.)
#  col 27  จังหวัด (ชส.)
#  col 28  รหัสไปรษณีย์ (ชส.)
#  col 29  โทรศัพท์ (ชส.)           → ไม่ import
#  col 30  อีเมล (ชส.)              → email
#  col 31  (ว่าง)
#
# ที่อยู่ที่ใช้ = col 24-28 รวมกัน ("ที่อยู่ที่ติดต่อได้ ชส.")
# email = col 30

def _parse_contact_row(row: list, row_num: int):
    errors = []

    tax_id = _clean_tax_id(row[1] if len(row) > 1 else None)
    if not tax_id or not tax_id.isdigit() or len(tax_id) not in (11, 13):
        errors.append(f"เลขประจำตัวผู้เสียภาษี '{tax_id}' ไม่ถูกต้อง")

    if errors:
        return None, {"row": row_num, "message": ", ".join(errors)}

    def _s(i):
        return str(row[i]).strip() if len(row) > i and row[i] is not None and str(row[i]).strip() else None

    # ที่อยู่ที่ติดต่อได้ (ชส.) = รวม col 24-28
    addr_parts = [_s(i) for i in range(24, 29) if _s(i)]
    address = " ".join(addr_parts) if addr_parts else None

    return {
        "tax_id":        tax_id,
        "operator_name": _s(5),
        "address":       address,
        "email":         _s(30),
    }, None


def parse_contact_excel(file_stream):
    """
    อ่านไฟล์ Excel ที่อยู่ผู้ประกอบการ (header แถว 5, ข้อมูลเริ่มแถว 6)
    คืน: (rows: list[dict], errors: list[dict])
    """
    try:
        _, data = _read_data_rows(file_stream, header_row=5)
    except Exception as e:
        return [], [{"row": 0, "message": f"เปิดไฟล์ไม่ได้: {e}"}]

    rows, errors = [], []
    for idx, row in enumerate(data, start=6):
        rec, err = _parse_contact_row(row, idx)
        if err:
            errors.append(err)
        else:
            rows.append(rec)
    return rows, errors


# ════════════════════════════════════════════════════════
# OPERATOR ACCOUNTS  →  operator_accounts
# ════════════════════════════════════════════════════════
# รองรับ 2 รูปแบบไฟล์:
#  1) template อย่างง่าย — header แถว 1: tax_id | operator_name | email (ติดกัน)
#  2) รายงานจริงของ กสทช. — header อยู่แถวอื่น คอลัมน์ไม่ติดกัน
#     (เช่น "เลขประจำตัวผู้เสียภาษี" อยู่คอลัมน์ B, "อีเมล" อยู่คอลัมน์ X)
# ตำแหน่งคอลัมน์จึงหาจาก "ชื่อ header" แทนตำแหน่งตายตัว

_TAX_ID_HEADER_ALIASES = {"tax_id", "เลขประจำตัวผู้เสียภาษี", "เลขที่ผู้เสียภาษี"}
_OPERATOR_NAME_HEADER_ALIASES = {
    "operator_name", "ชื่อผู้ประกอบการ", "รายชื่อผู้ประกอบการ", "รายชื่อผู้รับใบอนุญาต"
}
_EMAIL_HEADER_ALIASES = {"email", "อีเมล", "อีเมล์"}


def _is_greenish_fill(cell):
    """
    True ถ้าเซลล์ header ถูกไฮไลต์สีเขียว (ใช้แยกคอลัมน์ "อีเมล" ที่ถูกต้อง
    เวลาไฟล์มีคอลัมน์ชื่อ "อีเมล" ซ้ำหลายคอลัมน์ — กสทช. ใช้สีไฮไลต์แยกกลุ่ม
    ที่อยู่แต่ละบล็อก เช่น สีชมพู = ที่อยู่ออกใบเสร็จ, สีเขียว = ที่อยู่ติดต่อได้จริง)
    """
    fill = cell.fill
    if fill is None or fill.patternType != "solid":
        return False
    fg = fill.fgColor
    if fg is None or fg.type != "rgb" or not fg.rgb or len(fg.rgb) != 8:
        return False
    try:
        r, g, b = int(fg.rgb[2:4], 16), int(fg.rgb[4:6], 16), int(fg.rgb[6:8], 16)
    except ValueError:
        return False
    return g > r + 15 and g > b + 15


def _find_operator_account_columns(ws, max_scan_rows=10):
    """
    สแกนแถวแรกๆ ของไฟล์หาแถวที่มี header ตรงกับชื่อคอลัมน์ tax_id, operator_name, email
    ถ้าคอลัมน์ "อีเมล" ปรากฏมากกว่า 1 คอลัมน์ในแถวเดียวกัน (ไฟล์รายงานจริงที่มี
    บล็อกที่อยู่หลายชุด) ให้เลือกคอลัมน์ที่ไฮไลต์สีเขียว ถ้าไม่มีสีเขียวเลย
    ให้ใช้คอลัมน์ขวาสุดที่ชื่อตรงกัน (fallback)
    คืน (header_row_idx (0-based), tax_col, name_col, email_col) หรือ None ถ้าไม่พบ
    """
    for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan_rows)):
        tax_col = name_col = None
        email_candidates = []   # [(col_idx, is_green)]
        for c_idx, cell in enumerate(row):
            text = str(cell.value).strip().lower() if cell.value is not None else ""
            if tax_col is None and text in _TAX_ID_HEADER_ALIASES:
                tax_col = c_idx
            elif name_col is None and text in _OPERATOR_NAME_HEADER_ALIASES:
                name_col = c_idx
            elif text in _EMAIL_HEADER_ALIASES:
                email_candidates.append((c_idx, _is_greenish_fill(cell)))
        if tax_col is not None and name_col is not None and email_candidates:
            green_cols = [c for c, is_green in email_candidates if is_green]
            email_col = green_cols[-1] if green_cols else email_candidates[-1][0]
            return r_idx, tax_col, name_col, email_col
    return None


def _parse_operator_account_row(row, excel_row_num, seen_tax_ids, tax_col=0, name_col=1, email_col=2):
    errs = []

    tax_id = _clean_tax_id(row[tax_col] if len(row) > tax_col else None)
    if not tax_id or not tax_id.isdigit() or len(tax_id) != 13:
        errs.append(f"เลขผู้เสียภาษี '{tax_id}' ต้องเป็นตัวเลข 13 หลัก")
    elif tax_id in seen_tax_ids:
        errs.append(f"เลขผู้เสียภาษี '{tax_id}' ซ้ำในไฟล์")

    operator_name = str(row[name_col]).strip() if len(row) > name_col and row[name_col] is not None else ""
    if not operator_name:
        errs.append("ชื่อผู้ประกอบการว่าง")

    email = str(row[email_col]).strip() if len(row) > email_col and row[email_col] is not None else ""
    if not email or not EMAIL_RE.match(email):
        errs.append(f"email '{email}' รูปแบบไม่ถูกต้อง")

    if errs:
        return None, {"row": excel_row_num, "message": ", ".join(errs)}

    seen_tax_ids.add(tax_id)
    return {"tax_id": tax_id, "operator_name": operator_name, "email": email}, None


def parse_operator_account_excel(file_stream):
    """
    อ่านไฟล์ Excel บัญชีผู้ประกอบการ — หาตำแหน่งคอลัมน์ tax_id/operator_name/email
    จาก header (ดู _find_operator_account_columns) รองรับทั้ง template อย่างง่าย
    และรายงานจริงที่ header อยู่แถวอื่น
    คืน: (rows: list[dict], errors: list[dict])
    """
    if isinstance(file_stream, (bytes, bytearray)):
        file_stream = io.BytesIO(file_stream)
    try:
        wb = openpyxl.load_workbook(file_stream, data_only=True)
        ws = wb.active
        found = _find_operator_account_columns(ws)
        all_rows = list(ws.iter_rows(values_only=True))
        wb.close()
    except Exception as e:
        return [], [{"row": 0, "message": f"เปิดไฟล์ไม่ได้: {e}"}]

    if not found:
        return [], [{
            "row": 0,
            "message": "ไม่พบคอลัมน์ที่จำเป็น (เลขผู้เสียภาษี, ชื่อผู้ประกอบการ, อีเมล) ในไฟล์"
        }]
    header_row_idx, tax_col, name_col, email_col = found

    data = [
        list(r) for r in all_rows[header_row_idx + 1:]
        if any(c is not None and str(c).strip() for c in r)
    ]

    rows, errors = [], []
    seen_tax_ids = set()
    for idx, row in enumerate(data, start=header_row_idx + 2):   # Excel row number จริง
        rec, err = _parse_operator_account_row(row, idx, seen_tax_ids, tax_col, name_col, email_col)
        if err:
            errors.append(err)
        else:
            rows.append(rec)
    return rows, errors


def import_operator_accounts(db, rows):
    """
    Upsert operator_accounts  unique: tax_id
    - มีอยู่แล้ว → UPDATE operator_name, email
    - ยังไม่มี   → INSERT
    คืน: {"inserted": n, "updated": n, "snapshot": [...]}
    snapshot เก็บค่าก่อนนำเข้าของแต่ละแถว (old=None ถ้าเป็นแถวใหม่) สำหรับ rollback
    """
    inserted = updated = 0
    snapshot = []
    cur = db.cursor()
    for row in rows:
        cur.execute("SELECT * FROM operator_accounts WHERE tax_id=%s", (row["tax_id"],))
        existing = cur.fetchone()
        key = {"tax_id": row["tax_id"]}
        if existing:
            snapshot.append({"table": "operator_accounts", "key": key, "old": existing})
            cur.execute(
                "UPDATE operator_accounts SET operator_name=%s, email=%s WHERE tax_id=%s",
                (row["operator_name"], row["email"], row["tax_id"])
            )
            updated += 1
        else:
            snapshot.append({"table": "operator_accounts", "key": key, "old": None})
            cur.execute(
                "INSERT INTO operator_accounts (tax_id, operator_name, email) VALUES (%s, %s, %s)",
                (row["tax_id"], row["operator_name"], row["email"])
            )
            inserted += 1
    db.commit()
    cur.close()
    return {"inserted": inserted, "updated": updated, "snapshot": snapshot}


def import_contacts(db, rows):
    """
    อัปเดต address + email ใน taxpayer_master
    ใช้ tax_id เป็น key — อัปเดตทุก fiscal_year ของ tax_id นั้น
    คืน: {"inserted": 0, "updated": n, "not_found": n}
    """
    updated = not_found = 0
    cur = db.cursor()

    for row in rows:
        cur.execute(
            "SELECT COUNT(*) as cnt FROM taxpayer_master WHERE tax_id = %s",
            (row["tax_id"],)
        )
        result = cur.fetchone()
        cnt = result["cnt"] if isinstance(result, dict) else result[0]

        if cnt == 0:
            not_found += 1
            continue

        set_parts, params = [], []
        if row.get("address") is not None:
            set_parts.append("address = %s")
            params.append(row["address"])
        if row.get("email") is not None:
            set_parts.append("email = %s")
            params.append(row["email"])

        if not set_parts:
            continue

        params.append(row["tax_id"])
        cur.execute(
            f"UPDATE taxpayer_master SET {', '.join(set_parts)} WHERE tax_id = %s",
            params
        )
        updated += 1

    db.commit()
    cur.close()
    return {"inserted": 0, "updated": updated, "not_found": not_found}