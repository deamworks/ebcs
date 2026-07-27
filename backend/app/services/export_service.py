# ════════════════════════════════════════════════════════
# services/export_service.py
# สร้างรายงาน Excel 3 ฉบับ
#
# 1. รายงานข้อมูลผู้ประกอบการ
# 2. รายงานข้อมูลใบอนุญาต
# 3. รายงานชำระเงินกองทุน
#
# วันที่ใช้รูปแบบไทย: DD/MM/YYYY (พ.ศ.)
# ไม่มีแถวสรุปรวม
# ════════════════════════════════════════════════════════

import io
from datetime import datetime

import openpyxl
from openpyxl.styles import (
    Font, Alignment, Border, Side, PatternFill
)


# ════════════════════════════════════════════════════════
# Style กลาง
# ════════════════════════════════════════════════════════

FONT_NAME   = "TH SarabunPSK"
HEADER_BG   = "FFFFFF"   # พื้นหลังขาว
HEADER_FG   = "FF000000"   # ตัวอักษรดำ
ALT_ROW_BG  = "FFFFFF"   # พื้นหลังขาว (ไม่สลับสีแถว)

def hfont(size=11):
    return Font(name=FONT_NAME, size=size, bold=True, color=HEADER_FG)

def bfont(size=11, bold=False):
    return Font(name=FONT_NAME, size=size, bold=bold)

def hfill(color=None):
    if not color:
        return PatternFill(fill_type=None)
    return PatternFill("solid", fgColor=color)

def thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def center():
    return Alignment(horizontal="center", vertical="center",
                     wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center",
                     wrap_text=True)

def right():
    return Alignment(horizontal="right", vertical="center")


# ════════════════════════════════════════════════════════
# Helper
# ════════════════════════════════════════════════════════

def to_thai_date(val):
    """
    แปลงวันที่เป็นรูปแบบไทย DD/MM/YYYY (พ.ศ.)
    รองรับ datetime object และ string YYYY-MM-DD
    """
    if val is None:
        return ""

    if isinstance(val, datetime):
        y_th = val.year + 543
        return f"{val.day:02d}/{val.month:02d}/{y_th}"

    val = str(val).strip()
    if not val or val == "None":
        return ""

    try:
        parts = val.split("-")
        y_ce  = int(parts[0])
        m     = int(parts[1])
        d     = int(parts[2][:2])
        y_th  = y_ce + 543
        return f"{d:02d}/{m:02d}/{y_th}"
    except (ValueError, IndexError):
        return val


def fmt_num(val):
    """แปลงตัวเลขเป็น string มี comma"""
    try:
        return f"{float(val or 0):,.2f}"
    except (TypeError, ValueError):
        return "0.00"


def set_header_row(ws, headers, row=1, fill_color=None, col_styles=None):
    """
    สร้างแถว header
    fill_color=None → ไม่มีพื้นหลัง (ทุกคอลัมน์ ถ้าไม่ระบุ col_styles)
    fill_color="FFFF00" → พื้นเหลืองทุกคอลัมน์
    col_styles → list ของ (fill_color, font_color) ต่อคอลัมน์ (1-indexed
                 ตามลำดับ headers) ใช้เมื่อแต่ละคอลัมน์มีสีต่างกัน
                 เช่น รายงานผู้ประกอบการที่มีสีแยกกลุ่มคอลัมน์แจ้งเตือน
    """
    for col, (text, width) in enumerate(headers, start=1):
        this_fill = fill_color
        this_font_color = HEADER_FG
        if col_styles and col <= len(col_styles):
            style_fill, style_font = col_styles[col - 1]
            this_fill = style_fill
            this_font_color = style_font or HEADER_FG

        cell = ws.cell(row=row, column=col, value=text)
        # strip FF prefix จาก ARGB → RGB (openpyxl Font รับ 6 หลัก)
        fc = this_font_color or "000000"
        # ใช้ ARGB 8 หลักตรงๆ — openpyxl ต้องการ FF นำหน้า
        cell.font      = Font(name=FONT_NAME, size=11, bold=True, color=fc)
        cell.fill      = hfill(this_fill)
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border    = thin_border()
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(col)
        ].width = width
    ws.row_dimensions[row].height = 25


def set_data_cell(cell, value, row_num,
                  align="left", is_num=False):
    """ตั้งค่า cell ข้อมูล (ตัวอักษรดำ พื้นขาว ไม่มีแถบสีสลับแถว)"""
    cell.value     = value
    cell.font      = bfont()
    cell.border    = thin_border()
    cell.alignment = right() if is_num else (
        center() if align == "center" else left()
    )


THAI_MONTHS = [
    "", "มกราคม", "กุมภาพันธ์", "มีนาคม", "เมษายน",
    "พฤษภาคม", "มิถุนายน", "กรกฎาคม", "สิงหาคม",
    "กันยายน", "ตุลาคม", "พฤศจิกายน", "ธันวาคม",
]


def to_thai_date_long(val):
    """
    แปลงวันที่เป็นรูปแบบไทยเต็ม "D เดือน พ.ศ." เช่น 16 กรกฎาคม 2569
    รองรับ datetime object และ string YYYY-MM-DD
    ถ้าแปลงไม่ได้จะคืนค่าตามที่รับมา
    """
    if val is None:
        return ""

    if isinstance(val, datetime):
        y_th = val.year + 543
        return f"{val.day} {THAI_MONTHS[val.month]} {y_th}"

    val = str(val).strip()
    if not val or val == "None":
        return ""

    try:
        parts = val.split("-")
        y_ce = int(parts[0])
        m    = int(parts[1])
        d    = int(parts[2][:2])
        y_th = y_ce + 543
        return f"{d} {THAI_MONTHS[m]} {y_th}"
    except (ValueError, IndexError):
        return val


def set_title(ws, title, max_col, report_date=None):
    """
    สร้างแถว title ตามรูปแบบรายงานจริงของ กสทช.:
    แถว 1 เว้นว่าง, แถว 2 = ชื่อรายงาน + วันที่ข้อมูล (merge, ขึ้นบรรทัดใหม่),
    แถว 3 = header, แถว 4 เป็นต้นไป = ข้อมูล

    report_date: วันที่ที่จะแสดงใน "ข้อมูล ณ ..."
                 รับได้เป็น datetime, string YYYY-MM-DD, หรือ None
                 (None = ใช้วันที่ปัจจุบันโดยอัตโนมัติ)
                 รูปแบบที่แสดง: "16 กรกฎาคม 2569" (วัน เดือนไทยเต็ม พ.ศ.)
    """
    ws.row_dimensions[1].height = 10  # แถว 1 เว้นว่างตามไฟล์ต้นฉบับ

    ws.merge_cells(
        start_row=2, start_column=1,
        end_row=2,   end_column=max_col
    )

    if report_date is None:
        report_date = datetime.now()
    date_text = to_thai_date_long(report_date)

    title_text = f"{title}\nข้อมูล ณ {date_text}"
    cell           = ws.cell(row=2, column=1, value=title_text)
    cell.font      = Font(name=FONT_NAME, size=14,
                          bold=True, color="FF000000")
    cell.alignment = center()
    ws.row_dimensions[2].height = 40


def make_output(wb):
    """บันทึก workbook เป็น BytesIO"""
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ════════════════════════════════════════════════════════
# รายงาน 1: ข้อมูลผู้ประกอบการ
# ════════════════════════════════════════════════════════

# สีต่อคอลัมน์ของ header รายงานผู้ประกอบการ (ตรงตามไฟล์ต้นฉบับ กสทช.)
# (fill_color, font_color) ต่อคอลัมน์ที่ 1-42 ตามลำดับ
TAXPAYER_HEADER_COLORS = (
    [("FFFF00", "FF000000")] * 13 +      # 1-13   เหลือง
    [(None,     "FFFF0000")] * 3  +      # 14-16  ไม่มีพื้น ตัวอักษรแดง
    [(None,     "FF00B0F0")]     +       # 17     ไม่มีพื้น ตัวอักษรฟ้า
    [(None,     "FF000000")]     +       # 18     ไม่มีพื้น ตัวอักษรดำ
    [("DAE3F3", "FF000000")] * 7  +      # 19-25  ฟ้าอ่อน
    [("FBE5D6", "FF000000")] * 7  +      # 26-32  ส้มอ่อน
    [("E2EFDA", "FF000000")] * 7  +      # 33-39  เขียวอ่อน
    [("00B0F0", "FF000000")] * 3         # 40-42  ฟ้า
)


def export_taxpayer_report(db, year=None, year_from=None, year_to=None, report_date=None):
    """
    รายงานข้อมูลผู้ประกอบการ - ตรงตามรูปแบบไฟล์ราชการ (กสทช.):
    "รายงานข้อมูลผู้ประกอบการที่นำส่งเงินรายปีเข้ากองทุน..."

    คอลัมน์แรก 5 คอลัมน์ดึงจาก taxpayer_master โดยตรง
    (ลำดับ / เลขที่ผู้เสียภาษี / ผู้ประกอบการ / รอบระยะเวลาบัญชี
     / วันครบกำหนดชำระ) ส่วนคอลัมน์ที่เหลือเป็นคอลัมน์ติดตามการแจ้งเตือน
     ทางไปรษณีย์ตามแบบฟอร์มต้นฉบับ ซึ่งยังไม่มีข้อมูลใน schema ปัจจุบัน
     จึงปล่อยว่างไว้เพื่อให้โครงสร้างไฟล์ตรงกับต้นฉบับ

    - year_from/year_to: กรองช่วงปีบัญชี (ใช้แทน year เดี่ยวถ้าส่งมา)
    """
    conditions = []
    params     = []

    if year_from:
        conditions.append("fiscal_year >= %s")
        params.append(year_from)
    if year_to:
        conditions.append("fiscal_year <= %s")
        params.append(year_to)
    if not year_from and not year_to and year:
        conditions.append("fiscal_year = %s")
        params.append(year)

    where = "WHERE " + " AND ".join(conditions) if conditions else ""

    with db.cursor() as cur:
        cur.execute(f"""
            SELECT
                tax_id, operator_name, ref_no,
                fiscal_year, period_start, period_end,
                due_date
            FROM taxpayer_master
            {where}
            ORDER BY fiscal_year DESC, operator_name
        """, params)
        rows = cur.fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "รายงานข้อมูลผู้ประกอบการ"

    headers = [
        ("ลำดับ",                              6),
        ("เลขที่ผู้เสียภาษี",                  18),
        ("ผู้ประกอบการ",                       35),
        ("รอบระยะเวลาบัญชี",                   20),
        ("วันครบกำหนดชำระ",                    14),
        ("ที่อยู่ที่ติดต่อได้(ชส.)",           25),
        ("ตำบล",                               12),
        ("อำเภอ",                              12),
        ("จังหวัด",                            12),
        ("รหัสไปรษณีย์",                       10),
        ("คำนำหน้า",                           10),
        ("ตัวย่อ1",                            10),
        ("ตัวย่อ2",                            10),
        ("วันสุดท้ายที่ครบรอบระยะเวลาบัญชี \n(ออกแจ้ง 150 วัน)", 16),
        ("ครบกำหนดนำส่ง - เริ่มส่ง\nเตือน1 (+150วัน)",           16),
        ("ครบกำหนดนำส่ง - เริ่มส่ง\nเตือน2 (+37วัน หลังออกเตือน1)", 16),
        ("สถานะกิจการ \nณ เตือน 2",            14),
        ("สถานะอัพเดท ",                       12),
        ("เลขหนังสือแจ้งก่อนครบ 150 วัน",       16),
        ("ลงวันที่",                           12),
        ("เลข EMS แจ้ง 150 วัน(ครั้งที่ 1)",    16),
        ("เลขติดตามสถานะ\nใบตอบรับ",           14),
        ("ได้รับไปรษณีย์+ส่งใบฟ้าให้ ชส.\n(แจ้ง 150 วัน)", 14),
        ("ไปรษณีย์ตีกลับ\n(แจ้ง 150 วัน)",     14),
        ("ได้รับไปรษณีย์+ไม่มีใบฟ้าให้ ชส.\n(แจ้ง 150 วัน)", 14),
        ("เลขหนังสือแจ้งเตือน 1",               16),
        ("ลงวันที่",                           12),
        ("เลข EMS ส่งเตือน 1",                  16),
        ("เลขติดตามสถานะ\nใบตอบรับ",           14),
        ("ได้รับไปรษณีย์และส่งใบฟ้าให้ ชส. \n(เตือน 1) ", 14),
        ("ไปรษณีย์ตีกลับ(เตือน 1)",             14),
        ("ได้รับไปรษณีย์+ไม่มีใบฟ้าให้ ชส.(เตือน 1) ", 14),
        ("เลขหนังสือแจ้งเตือน 2",               16),
        ("ลงวันที่",                           12),
        ("เลข EMS ส่งเตือน 2 (ครั้งที่1)",      16),
        ("เลขติดตามสถานะ\nใบตอบรับ",           14),
        ("ได้รับไปรษณีย์และส่งใบฟ้าให้ ชส. \n(เตือน 2) ", 14),
        ("ไปรษณีย์ตีกลับ(เตือน 2)",             14),
        ("ได้รับไปรษณีย์+ไม่มีใบฟ้าให้ ชส.(เตือน 2) ", 14),
        ("คำนำหน้า",                           10),
        ("ตัวย่อ1",                            10),
        ("ตัวย่อ2",                            10),
    ]

    # ── title รายงาน 2 ผู้ประกอบการ ────────────────────────
    # row2: ชื่อรายงาน + วันที่ 2 บรรทัดใน cell เดียว — left, size=14, bold
    # [FIX] ปีที่รับมาเป็น พ.ศ. อยู่แล้วเสมอ ไม่ต้องบวก 543 ซ้ำ — รองรับช่วงปีด้วย
    if year_from and year_to:
        year_be_tp = f"{year_from} - {year_to}"
    elif year_from:
        year_be_tp = f"{year_from} เป็นต้นไป"
    elif year_to:
        year_be_tp = f"ถึง {year_to}"
    else:
        year_be_tp = year or None
    now_tp       = report_date if isinstance(report_date, datetime) else datetime.now()
    date_text_tp = to_thai_date_long(now_tp)

    ws.row_dimensions[1].height = 10
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    tp_title = (
        "รายงานข้อมูลผู้ประกอบการที่นำส่งเงินรายปีเข้ากองทุนวิจัยและพัฒนา"
        "กิจการกระจายเสียง กิจการโทรทัศน์ และกิจการโทรคมนาคม เพื่อประโยชน์สาธารณะ"
        + (f" สำหรับรอบระยะเวลาบัญชีปี {year_be_tp}" if year_be_tp else "")
        + f"\nข้อมูล ณ {date_text_tp}"
    )
    c2 = ws.cell(row=2, column=1, value=tp_title)
    c2.font      = Font(name=FONT_NAME, size=14, bold=True, color="FF000000")
    c2.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 50
    set_header_row(ws, headers, row=3, col_styles=TAXPAYER_HEADER_COLORS)

    for i, row in enumerate(rows, start=1):
        r = i + 3
        # 5 คอลัมน์แรกมีข้อมูลจริง ที่เหลือปล่อยว่าง (ยังไม่มีใน schema)
        data = [
            i,
            row["tax_id"],
            row["operator_name"],
            f"{to_thai_date(row['period_start'])} - "
            f"{to_thai_date(row['period_end'])}",
            to_thai_date(row["due_date"]),
        ] + [""] * (len(headers) - 5)

        for c, val in enumerate(data, start=1):
            set_data_cell(
                ws.cell(row=r, column=c, value=val),
                val, i,
                align="center" if c == 1 else "left",
                is_num=False
            )
        ws.row_dimensions[r].height = 20

    return make_output(wb)


# ════════════════════════════════════════════════════════
# รายงาน 2: ข้อมูลใบอนุญาต
# ════════════════════════════════════════════════════════

def export_licensee_report(db, year=None, year_from=None, year_to=None, status=None, report_date=None):
    """
    รายงานข้อมูลใบอนุญาต - ตรงตามรูปแบบไฟล์ราชการ (กสทช.):
    "ผลการจัดเก็บเงินรายปีเข้ากองทุนวิจัยและพัฒนา..."

    คอลัมน์การเงิน (รายได้จากใบอนุญาต / ค่าลดหย่อน / เงินนำส่งเข้ากองทุน
    / ภาษีมูลค่าเพิ่ม / เงินเพิ่ม / จำนวนเงินรายปีนำส่งเข้ากองทุนสุทธิ) ยังไม่มีใน
    licensee_master schema ปัจจุบัน จึงปล่อยว่างไว้เพื่อให้โครงสร้างคอลัมน์ตรงกับ
    ต้นฉบับ — "รหัสอ้างอิง" join มาจาก taxpayer_master, "สถานะ" (แนบเอกสาร) derive
    จากใบยื่นแบบของ tax_id/ปีเดียวกัน, "ประเภท"/"รอบ" ยังไม่มีแนวคิดแยกในระบบ
    จึงใช้ค่า "ปกติ" คงที่เหมือนรายงานชำระเงินกองทุน

    - year_from/year_to: กรองช่วงปีบัญชี (ใช้แทน year เดี่ยวถ้าส่งมา)
    """
    conditions = []
    params     = []

    if year_from:
        conditions.append("lm.fiscal_year >= %s")
        params.append(year_from)
    if year_to:
        conditions.append("lm.fiscal_year <= %s")
        params.append(year_to)
    if not year_from and not year_to and year:
        conditions.append("lm.fiscal_year = %s")
        params.append(year)

    if status:
        conditions.append("lm.license_status = %s")
        params.append(status)

    where = "WHERE " + " AND ".join(conditions) if conditions else ""

    status_labels = {
        "active":    "ปกติ",
        "ended":     "สิ้นสุดระยะเวลา",
        "cancelled": "ยกเลิก",
        "revoked":   "เพิกถอน",
    }

    with db.cursor() as cur:
        # [FIX] "รหัสอ้างอิง" (ref_no) ไม่เคยดึงมาเลย — ref_no อยู่ใน
        # taxpayer_master (key ด้วย tax_id + fiscal_year) ต้อง join เอา
        cur.execute(f"""
            SELECT
                lm.license_no, lm.tax_id, lm.company_name,
                lm.licensee_type AS license_type, lm.license_status,
                lm.start_date AS license_start, lm.end_date AS license_end,
                lm.fiscal_year, tm.ref_no
            FROM licensee_master lm
            LEFT JOIN taxpayer_master tm
                   ON tm.tax_id = lm.tax_id AND tm.fiscal_year = lm.fiscal_year
            {where}
            ORDER BY lm.license_no
        """, params)
        rows = cur.fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "รายงานข้อมูลใบอนุญาต"

    headers = [
        ("ลำดับ",                      6),
        ("ปี",                         8),
        ("รหัสอ้างอิง",                 14),
        ("ประเภท",                     10),
        ("รอบ",                        12),
        ("สถานะ",                      12),
        ("เลขประจำตัวผู้เสียภาษี",     16),
        ("รายชื่อผู้รับใบอนุญาต",       35),
        ("จำนวนใบอนุญาต",              10),
        ("เลขที่ใบอนุญาต",             18),
        ("ประเภทใบอนุญาต",              14),
        ("วันที่เริ่มต้น",              14),
        ("วันที่สิ้นสุด",               14),
        ("สถานะใบอนุญาต",              14),
        ("รายได้จากใบอนุญาต",           16),
        ("ค่าลดหย่อน",                  14),
        ("เงินนำส่งเข้ากองทุน",         16),
        ("ภาษีมูลค่าเพิ่ม",             14),
        ("เงินเพิ่ม",                   12),
        ("จำนวนเงินรายปีนำส่งเข้ากองทุนสุทธิ", 18),
    ]

    # ── title รายงาน 3 ใบอนุญาต ────────────────────────────
    # row2: 3 บรรทัดใน cell เดียว — left, size=11, ไม่ bold
    if year_from and year_to:
        yr_range = f"{year_from}–{year_to}"
    elif year_from:
        yr_range = f"{year_from} เป็นต้นไป"
    elif year_to:
        yr_range = f"ถึง {year_to}"
    else:
        yr_range = f"{year}–{year}" if year else ""
    now_lic      = report_date if isinstance(report_date, datetime) else datetime.now()
    date_text_lic = to_thai_date_long(now_lic)

    ws.row_dimensions[1].height = 10
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    lic_title = (
        "ผลการจัดเก็บเงินรายปีเข้ากองทุนวิจัยและพัฒนากิจการกระจายเสียง "
        "กิจการโทรทัศน์ และกิจการโทรคมนาคม เพื่อประโยชน์สาธารณะ"
        + (f" สำหรับรอบระยะเวลาบัญชีปี {yr_range}" if yr_range else "")
        + f"\nจัดเก็บในปี {yr_range} ซึ่งเป็นรายได้ที่นำมาคำนวณตามอัตราการนำส่งเงินรายปีเข้ากองทุนวิจัยและพัฒนาฯ"
        + f"\nตามประกาศ กสทช. เรื่อง หลักเกณฑ์และวิธีการนำส่งเงินรายปีเข้ากองทุนวิจัยและพัฒนากิจการกระจายเสียง "
        + "กิจการโทรทัศน์ และกิจการโทรคมนาคม เพื่อประโยชน์สาธารณะ "
        + "ของผู้รับใบอนุญาตประกอบกิจการกระจายเสียงหรือกิจการโทรทัศน์ (ฉบับที่ 2)"
        + f"ข้อมูล ณ {date_text_lic}"
    )
    c2 = ws.cell(row=2, column=1, value=lic_title)
    c2.font      = Font(name=FONT_NAME, size=11, bold=True, color="FF000000")
    c2.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 80
    set_header_row(ws, headers, row=3)

    # นับจำนวนใบอนุญาตต่อ tax_id เพื่อกรอกคอลัมน์ "จำนวนใบอนุญาต"
    license_counts = {}
    for row in rows:
        license_counts[row["tax_id"]] = (
            license_counts.get(row["tax_id"], 0) + 1
        )

    # [FIX] "สถานะ" (สถานะแนบเอกสาร) ไม่เคยดึงมาเลย — หาจากใบยื่นแบบ (submissions)
    # ของ tax_id + fiscal_year เดียวกัน คำนวณ pending_attach เหมือนรายงาน/หน้ารายการอื่น
    submission_status_map = {}
    with db.cursor() as cur:
        cur.execute("""
            SELECT s.tax_id, s.fiscal_year, s.status,
                   r.id AS receipt_id,
                   (SELECT COUNT(*) FROM document_attachments da
                    WHERE da.submission_id = s.id) AS attachment_count
            FROM submissions s
            LEFT JOIN receipt r ON r.submission_id = s.id
        """)
        for srow in cur.fetchall():
            st = "paid" if srow["receipt_id"] else srow["status"]
            if st == "pending_payment" and (srow["attachment_count"] or 0) < 3:
                st = "pending_attach"
            submission_status_map[(srow["tax_id"], srow["fiscal_year"])] = st

    submission_status_labels = {
        "draft":           "ร่าง",
        "pending_attach":  "รอแนบ",
        "pending_payment": "รอชำระเงิน",
        "paid":            "ชำระแล้ว",
    }

    for i, row in enumerate(rows, start=1):
        r = i + 3
        sub_status = submission_status_map.get((row["tax_id"], row["fiscal_year"]))
        data = [
            i,
            row["fiscal_year"] or "",
            row.get("ref_no") or "",
            "ปกติ",  # ประเภท (งวดนำส่ง) - ไม่มีข้อมูลแยกประเภทอื่นในระบบ ใช้ค่าเดียวกับรายงานชำระเงิน
            "ปกติ",  # รอบ - เช่นเดียวกัน
            submission_status_labels.get(sub_status, "") if sub_status else "",  # สถานะแนบเอกสาร
            row["tax_id"] or "",
            row["company_name"],
            license_counts.get(row["tax_id"], 1),
            row["license_no"],
            row["license_type"] or "",
            to_thai_date(row["license_start"]),
            to_thai_date(row["license_end"]),
            status_labels.get(
                row["license_status"], row["license_status"]
            ),
            "", "", "", "", "", "",  # คอลัมน์การเงิน - ยังไม่มีใน schema
        ]
        num_cols = {2, 9, 15, 16, 17, 18, 19, 20}
        for c, val in enumerate(data, start=1):
            set_data_cell(
                ws.cell(row=r, column=c, value=val),
                val, i,
                align="center" if c in {1, 2, 9} else "left",
                is_num=c in num_cols
            )
        ws.row_dimensions[r].height = 20

    return make_output(wb)



# ════════════════════════════════════════════════════════
# รายงาน 3: รายงานชำระเงินกองทุน (67 คอลัมน์)
# ตรงตามไฟล์ต้นฉบับ กสทช. ทุกสี / merge / width / height
#
# สีกลุ่ม header (row7+row8):
#  col 1-25   : ไม่มีพื้น (00000000) — merge row7:row8 แต่ละ col
#  col 26-30  : ไม่มีพื้น (sub row8 ไม่มีพื้นเช่นกัน)  "รายได้ตามงบแจกแจง"
#  col 31-35  : FFF1E0 (ส้มอ่อน)                       "ค่าธรรมเนียม"
#  col 36-43  : FFFFCC (เหลืองอ่อน)                    "ใบแจ้งหนี้"
#  col 44-47  : CCFFCC (เขียวอ่อน)                     "ใบเสร็จ"
#  col 48-52  : EEEEEE (เทาอ่อน)                       "ผลการสอบทาน"
#  col 53-56  : A5CDFF (ฟ้าอ่อน)                       "ผลต่างจากการสอบทาน"
#  col 57-63  : FFECF3 (ชมพูอ่อน)                      "ใบแจ้งหนี้/ใบเพิ่มหนี้"
#  col 64-67  : E3F3F0 (เขียวมิ้นต์)                   "ใบเสร็จ" (ชุดที่ 2)
#
# row2-6 height=21, row7 height=30, row8 height=36, data height=18
# ════════════════════════════════════════════════════════

# สีแต่ละกลุ่ม (ตรงกับไฟล์จริง)
_FILL = {
    "none":    "00000000",
    "orange":  "FFFFF1E0",
    "yellow":  "FFFFFFCC",
    "green":   "FFCCFFCC",
    "gray":    "FFEEEEEE",
    "blue":    "FFA5CDFF",
    "pink":    "FFFFECF3",
    "mint":    "FFE3F3F0",
}

# col → fill key (1-indexed)
def _col_fill(col):
    if col <= 25:  return "none"
    if col <= 30:  return "none"
    if col <= 35:  return "orange"
    if col <= 43:  return "yellow"
    if col <= 47:  return "green"
    if col <= 52:  return "gray"
    if col <= 56:  return "blue"
    if col <= 63:  return "pink"
    return "mint"

# col widths ตรงกับไฟล์จริงทุก col
_COL_WIDTHS = {
    1:13, 2:13, 3:13, 4:24, 5:17, 6:13, 7:18, 8:13, 9:13,
    10:13, 11:13, 12:13, 13:12, 14:13, 15:13, 16:13, 17:10,
    18:13, 19:12, 20:13, 21:13, 22:12, 23:13, 24:13, 25:12,
    26:13, 27:13, 28:13, 29:13, 30:13,
    31:13, 32:13, 33:13, 34:13, 35:13,
    36:13, 37:13, 38:13, 39:13, 40:11, 41:13, 42:13, 43:13,
    44:13, 45:13, 46:13, 47:13,
    48:13, 49:13, 50:13, 51:13, 52:13,
    53:13, 54:13, 55:13, 56:13,
    57:13, 58:13, 59:12, 60:13, 61:13, 62:13, 63:13,
    64:13, 65:13, 66:13, 67:13,
}


def export_payment_report(db, year=None, year_from=None, year_to=None, statuses=None, report_date=None):
    """
    Export รายงานชำระเงินกองทุน 67 คอลัมน์
    ตรงกับไฟล์ต้นฉบับ กสทช. ทุกสี / merge / width / height

    - year: ค่าเดียว (compat เดิม) — ใช้ได้ถ้าไม่ส่ง year_from/year_to
    - year_from/year_to: กรองช่วงปีบัญชี ใช้ค่าใดค่าหนึ่งหรือทั้งคู่ก็ได้
    - statuses: list ของ raw status ('draft'/'pending_attach'/'pending_payment'/'paid')
      กรองหลัง derive actual_status แล้ว (ดู pending_attach ด้านล่าง)
    """
    from openpyxl.utils import get_column_letter

    # ── Query ──────────────────────────────────────────────
    conditions, params = ["1=1"], []
    if year_from:
        conditions.append("s.fiscal_year >= %s")
        params.append(year_from)
    if year_to:
        conditions.append("s.fiscal_year <= %s")
        params.append(year_to)
    if not year_from and not year_to and year:
        conditions.append("s.fiscal_year = %s")
        params.append(year)
    where = "WHERE " + " AND ".join(conditions)

    with db.cursor() as cur:
        cur.execute(f"""
            SELECT
                s.id, s.tax_id, s.operator_name, s.ref_no,
                s.fiscal_year, s.period_start, s.period_end,
                s.due_date, s.status,
                s.total_income, s.deduction_amount,
                s.fund_amount, s.vat_amount, s.extra_amount, s.net_amount,
                s.submitted_at,
                CASE WHEN r.id IS NOT NULL THEN 'paid' ELSE s.status END AS actual_status,
                r.receipt_no, r.received_at, r.amount AS receipt_amount,
                inv.invoice_no, inv.issued_at, inv.amount AS invoice_amount,
                (SELECT COUNT(*) FROM licensee_master lm
                 WHERE lm.tax_id = s.tax_id) AS license_count,
                (SELECT COUNT(*) FROM document_attachments da
                 WHERE da.submission_id = s.id) AS attachment_count
            FROM submissions s
            LEFT JOIN receipt r   ON r.submission_id = s.id
            LEFT JOIN invoice inv ON inv.submission_id = s.id
            {where}
            GROUP BY s.id, r.id, inv.id
            ORDER BY s.fiscal_year DESC, s.operator_name
        """, params)
        rows = cur.fetchall()

    # [FIX] เดิม actual_status คำนวณแค่ paid/status ดิบ ไม่เคยเช็คเอกสารแนบ
    # บังคับ 3 อย่างครบไหม ทำให้รายงาน export ไม่ตรงกับสถานะ "รอแนบ" ที่แสดง
    # ในหน้ารายการแอดมิน (GET /admin/submissions) ให้คำนวณเหมือนกัน
    for row in rows:
        if row["actual_status"] == "pending_payment" and (row.get("attachment_count") or 0) < 3:
            row["actual_status"] = "pending_attach"

    if statuses:
        rows = [r for r in rows if r["actual_status"] in statuses]

    # ── Workbook ────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "รายงานชำระเงินกองทุน"

    TOTAL_COLS = 67
    LAST_COL   = get_column_letter(TOTAL_COLS)  # BO

    now = report_date if isinstance(report_date, datetime) else datetime.now()
    month_th = THAI_MONTHS[now.month]
    # [FIX] ปีที่รับมาทาง query param เป็น พ.ศ. อยู่แล้วเสมอ (ตรงกับที่เก็บใน
    # fiscal_year) ไม่ต้องบวก 543 ซ้ำ — และรองรับแสดงเป็นช่วงถ้าใช้ year_from/year_to
    if year_from and year_to:
        year_be = f"{year_from} - {year_to}"
    elif year_from:
        year_be = f"{year_from} เป็นต้นไป"
    elif year_to:
        year_be = f"ถึง {year_to}"
    else:
        year_be = year or ""
    round_str = "ปกติ, อื่นๆ"

    # ── ตัวช่วย style ────────────────────────────────────────
    THIN = Side(style="thin")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    def _pfill(hex_rgba):
        if not hex_rgba or hex_rgba == "00000000":
            return PatternFill(fill_type=None)
        return PatternFill("solid", fgColor=hex_rgba)

    def _cell_style(c, fill_key="none", bold=True, size=11,
                    halign="center", border=True):
        c.font      = Font(name=FONT_NAME, size=size, bold=bold, color="FF000000")
        c.fill      = _pfill(_FILL[fill_key])
        c.alignment = Alignment(horizontal=halign, vertical="center", wrap_text=True)
        if border:
            c.border = BORDER

    def _merge_border(r1, c1, r2, c2, fill_key="none", text=None):
        """
        merge cells แล้วใส่ border ทุก cell ในกลุ่ม
        openpyxl ใส่ border ให้แค่ top-left cell —
        ต้องวน set ทุก cell เพื่อให้เส้นเต็มทุกด้าน
        """
        ws.merge_cells(start_row=r1, start_column=c1,
                       end_row=r2,   end_column=c2)
        for row in range(r1, r2 + 1):
            for col in range(c1, c2 + 1):
                c = ws.cell(row=row, column=col)
                # border ทุก cell
                left_side   = THIN if col == c1 else Side(style=None)
                right_side  = THIN if col == c2 else Side(style=None)
                top_side    = THIN if row == r1 else Side(style=None)
                bottom_side = THIN if row == r2 else Side(style=None)
                c.border = Border(left=left_side, right=right_side,
                                  top=top_side, bottom=bottom_side)
                c.fill = _pfill(_FILL[fill_key])
                c.font = Font(name=FONT_NAME, size=11, bold=True, color="FF000000")
                c.alignment = Alignment(horizontal="center", vertical="center",
                                        wrap_text=True)
        # ใส่ข้อความที่ cell แรก
        if text is not None:
            ws.cell(r1, c1).value = text

    # ── set col widths ───────────────────────────────────────
    for col, w in _COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    # ── row 1: ว่าง ─────────────────────────────────────────
    ws.row_dimensions[1].height = 8

    # ── row 2-6: หัวรายงาน (size=14, bold, center, no fill)
    #  merge A:I เหมือนต้นฉบับ ──────────────────────────────
    # row2-6: หัวรายงาน — center, size=14, bold, merge A:BO (ทุกแถว)
    titles_r2_r6 = [
        "กองทุนวิจัยและพัฒนากิจการกระจายเสียง กิจการโทรทัศน์ และกิจการโทรคมนาคมเพื่อประโยชน์สาธารณะ",
        "รายงานข้อมูลการชำระเงินกองทุนประจำปี",
        f"ข้อมูลตามปีบัญชี {year_be}",
        f"รอบ : {round_str}",
        f"ข้อมูล ณ วันที่ {now.day} {month_th} {now.year + 543}",
    ]
    for r, text in enumerate(titles_r2_r6, start=2):
        ws.merge_cells(f"A{r}:{LAST_COL}{r}")
        c = ws.cell(row=r, column=1, value=text)
        c.font      = Font(name=FONT_NAME, size=14, bold=True, color="FF000000")
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[r].height = 21

    # ── row 7: header หลัก ──────────────────────────────────
    # col 1-25: merge row7:row8 ทีละ col ไม่มีพื้น
    single_headers = [
        (1,  "ลำดับ"),
        (2,  "ปีบัญชี"),
        (3,  "รหัสอ้างอิง"),
        (4,  "รายชื่อผู้รับใบอนุญาต"),
        (5,  "เลขประจำผู้เสียภาษี"),
        (6,  "สถานะกิจการ"),
        (7,  "รอบระยะเวลาบัญชี"),
        (8,  "ประเภท"),
        (9,  "รอบบัญชี"),
        (10, "วันที่เริ่มต้น\nรอบบัญชี"),
        (11, "วันที่สิ้นสุด\nรอบบัญชี"),
        (12, "วันที่ครบ\nกำหนดชำระ"),
        (13, "สถานะการชำระเงิน"),
        (14, "วันที่จ่ายชำระ"),
        (15, "จำนวนใบอนุญาต"),
        (16, "งบการเงินปี"),
        (17, "รายได้รวมตามงบการเงิน"),
        (18, "รายได้อื่นที่ไม่นำมาคำนวณ"),
        (19, "รายได้จากใบอนุญาต"),
        (20, "ค่าลดหย่อน"),
        (21, "รายได้หลังการลดหย่อน"),
        (22, "เงินนำส่งเข้ากองทุน"),
        (23, "ภาษีมูลค่าเพิ่ม"),
        (24, "เงินเพิ่ม"),
        (25, "จำนวนเงินรายปีนำส่ง\nเข้ากองทุนสุทธิ"),
    ]
    for col, text in single_headers:
        _merge_border(7, col, 8, col, fill_key="none", text=text)

    # col 26-30: "รายได้ตามงบการเงินแจกแจงตามประเภท" (row7 merge Z:AD)
    _merge_border(7, 26, 7, 30, fill_key="none", text="รายได้ตามงบการเงินแจกแจงตามประเภท")
    for col, sub in [(26,"รายได้ radio"),(27,"รายได้ service"),
                     (28,"รายได้ network"),(29,"รายได้ facility"),(30,"รวมรายได้")]:
        c8 = ws.cell(row=8, column=col, value=sub)
        _cell_style(c8, fill_key="none")

    # col 31-35: "ค่าธรรมเนียมนำส่งเข้ากองทุนวิจัยฯ ตามประกาศฯ" (AE7:AI7) สีส้มอ่อน
    _merge_border(7, 31, 7, 35, fill_key="orange", text="ค่าธรรมเนียมนำส่งเข้ากองทุนวิจัยฯ ตามประกาศฯ")
    for col, sub in [(31,"fee radio"),(32,"fee service"),
                     (33,"fee network"),(34,"fee facility"),(35,"รวมค่าธรรมเนียม")]:
        c8 = ws.cell(row=8, column=col, value=sub)
        _cell_style(c8, fill_key="orange")

    # col 36-43: "ใบแจ้งหนี้" (AJ7:AQ7) สีเหลืองอ่อน
    _merge_border(7, 36, 7, 43, fill_key="yellow", text="ใบแจ้งหนี้")
    for col, sub in [
        (36,"Ref.1"),(37,"Ref.2"),(38,"วันที่"),(39,"เลขที่"),
        (40,"จำนวนเงิน\nก่อนภาษีมูลค่าเพิ่ม"),(41,"ภาษีมูลค่าเพิ่ม"),
        (42,"เงินเพิ่ม"),(43,"จำนวนเงิน\nรวมทั้งสิ้น")]:
        c8 = ws.cell(row=8, column=col, value=sub)
        _cell_style(c8, fill_key="yellow")

    # col 44-47: "ใบเสร็จ" (AR7:AU7) สีเขียวอ่อน
    _merge_border(7, 44, 7, 47, fill_key="green", text="ใบเสร็จ")
    for col, sub in [(44,"วันที่"),(45,"เลขที่"),
                     (46,"จำนวนเงิน"),(47,"จำนวนเงิน\nรวมทั้งสิ้น")]:
        c8 = ws.cell(row=8, column=col, value=sub)
        _cell_style(c8, fill_key="green")

    # col 48-52: "ผลการสอบทาน" (AV7:AZ7) สีเทาอ่อน
    _merge_border(7, 48, 7, 52, fill_key="gray", text="ผลการสอบทาน")
    for col, sub in [
        (48,"สอบทาน\nครั้งที่"),(49,"เงินนำส่งเข้ากองทุน"),
        (50,"ภาษีมูลค่าเพิ่ม"),(51,"เงินเพิ่ม"),(52,"จำนวนเงิน\nรวมทั้งสิ้น")]:
        c8 = ws.cell(row=8, column=col, value=sub)
        _cell_style(c8, fill_key="gray")

    # col 53-56: "ผลต่างจากการสอบทาน" (BA7:BD7) สีฟ้าอ่อน
    _merge_border(7, 53, 7, 56, fill_key="blue", text="ผลต่างจากการสอบทาน")
    for col, sub in [
        (53,"เงินนำส่งเข้ากองทุน"),(54,"ภาษีมูลค่าเพิ่ม"),
        (55,"เงินเพิ่ม"),(56,"จำนวนเงิน\nรวมทั้งสิ้น")]:
        c8 = ws.cell(row=8, column=col, value=sub)
        _cell_style(c8, fill_key="blue")

    # col 57-63: "ใบแจ้งหนี้/ใบเพิ่มหนี้" (BE7:BK7) สีชมพูอ่อน
    _merge_border(7, 57, 7, 63, fill_key="pink", text="ใบแจ้งหนี้/ใบเพิ่มหนี้")
    for col, sub in [
        (57,"วันที่ใบแจ้งหนี้"),(58,"เลขที่ใบแจ้งหนี้"),
        (59,"จำนวนเงิน\nก่อนภาษีมูลค่าเพิ่ม"),(60,"ภาษีมูลค่าเพิ่ม"),
        (61,"เงินเพิ่ม"),(62,"ค่าปรับเงินเพิ่ม"),(63,"จำนวนเงิน\nรวมทั้งสิ้น")]:
        c8 = ws.cell(row=8, column=col, value=sub)
        _cell_style(c8, fill_key="pink")

    # col 64-67: "ใบเสร็จ" (BL7:BO7) สีเขียวมิ้นต์
    _merge_border(7, 64, 7, 67, fill_key="mint", text="ใบเสร็จ")
    for col, sub in [(64,"วันที่ใบเสร็จ"),(65,"เลขที่ใบเสร็จ"),
                     (66,"จำนวนเงิน"),(67,"จำนวนเงิน\nรวมทั้งสิ้น")]:
        c8 = ws.cell(row=8, column=col, value=sub)
        _cell_style(c8, fill_key="mint")

    ws.row_dimensions[7].height = 30
    ws.row_dimensions[8].height = 36

    # ── row 9+: ข้อมูล ─────────────────────────────────────
    status_map = {
        "draft":           "ร่าง",
        "pending_attach":  "รอแนบ",
        "pending_payment": "รอชำระเงิน",
        "paid":            "ชำระแล้ว",
    }
    RIGHT_COLS  = set(range(17, 26)) | set(range(26, 36)) |                   {40,41,42,43,46,47,49,50,51,52,
                   53,54,55,56,59,60,61,62,63,66,67}
    CENTER_COLS = {1, 2, 5, 8, 9, 15, 16}

    m_short = ["","ม.ค.","ก.พ.","มี.ค.","เม.ย.","พ.ค.","มิ.ย.",
               "ก.ค.","ส.ค.","ก.ย.","ต.ค.","พ.ย.","ธ.ค."]

    for i, row in enumerate(rows, start=1):
        excel_row = i + 8

        # [FIX] fiscal_year เก็บเป็น พ.ศ. อยู่แล้ว (import_service.py แปลงไว้ตอน
        # นำเข้า) เดิมบวก 543 ซ้ำอีกรอบตรงนี้ ทำให้ปี 2569 กลายเป็น 3112
        fiscal_be = row["fiscal_year"] if row.get("fiscal_year") else ""
        fin_yr_be = fiscal_be

        # รอบระยะเวลา: "01 ม.ค. 68 - 31 ธ.ค. 68"
        period_str = ""
        ps = row.get("period_start")
        pe = row.get("period_end")
        if ps and pe and hasattr(ps, "month"):
            period_str = (f"{ps.day:02d} {m_short[ps.month]} "
                          f"{str(ps.year+543)[2:]} - "
                          f"{pe.day:02d} {m_short[pe.month]} "
                          f"{str(pe.year+543)[2:]}")
        elif ps and pe:
            period_str = f"{to_thai_date(ps)} - {to_thai_date(pe)}"

        pay_status  = status_map.get(row.get("actual_status",""), row.get("actual_status","") or "")
        inv_amount  = float(row.get("invoice_amount") or 0)
        net_inv     = inv_amount + float(row.get("vat_amount") or 0) + float(row.get("extra_amount") or 0)
        rcp_amount  = float(row.get("receipt_amount") or 0)

        # 67 ค่า
        values = [
            i,                                                   # 1
            fiscal_be,                                           # 2
            row.get("ref_no") or "",                             # 3
            row.get("operator_name") or "",                      # 4
            row.get("tax_id") or "",                             # 5
            "",                                                  # 6  สถานะกิจการ
            period_str,                                          # 7
            "ปกติ",                                              # 8
            "ปกติ",                                              # 9
            to_thai_date(row.get("period_start")),               # 10
            to_thai_date(row.get("period_end")),                 # 11
            to_thai_date(row.get("due_date")),                   # 12
            pay_status,                                          # 13
            to_thai_date(row.get("received_at")),                # 14
            int(row.get("license_count") or 0),                 # 15
            fin_yr_be,                                           # 16
            fmt_num(row.get("total_income")),                    # 17
            "0.00",                                              # 18
            fmt_num(row.get("total_income")),                    # 19
            fmt_num(row.get("deduction_amount")),                # 20
            fmt_num(float(row.get("total_income") or 0) -
                    float(row.get("deduction_amount") or 0)),    # 21
            fmt_num(row.get("fund_amount")),                     # 22
            fmt_num(row.get("vat_amount")),                      # 23
            fmt_num(row.get("extra_amount")),                    # 24
            fmt_num(row.get("net_amount")),                      # 25
            # 26-30 รายได้แจกแจง (ว่าง)
            "","","","","",
            # 31-35 ค่าธรรมเนียม (ว่าง)
            "","","","","",
            # 36-43 ใบแจ้งหนี้
            row.get("ref_no") or "",                             # 36 Ref.1
            "",                                                  # 37 Ref.2
            to_thai_date(row.get("issued_at")),                  # 38 วันที่
            row.get("invoice_no") or "",                         # 39 เลขที่
            fmt_num(inv_amount),                                 # 40
            fmt_num(row.get("vat_amount")),                      # 41
            fmt_num(row.get("extra_amount")),                    # 42
            fmt_num(net_inv),                                    # 43
            # 44-47 ใบเสร็จ
            to_thai_date(row.get("received_at")),                # 44
            row.get("receipt_no") or "",                         # 45
            fmt_num(rcp_amount),                                 # 46
            fmt_num(rcp_amount),                                 # 47
            # 48-52 ผลการสอบทาน (ว่าง)
            "","","","","",
            # 53-56 ผลต่าง (ว่าง)
            "","","","",
            # 57-63 ใบแจ้งหนี้/ใบเพิ่มหนี้ (ว่าง)
            "","","","","","","",
            # 64-67 ใบเสร็จ2 (ว่าง)
            "","","","",
        ]

        for col, val in enumerate(values, start=1):
            c = ws.cell(row=excel_row, column=col, value=val)
            c.font   = Font(name=FONT_NAME, size=11, bold=False, color="FF000000")
            c.border = BORDER
            c.alignment = Alignment(
                horizontal=("right"  if col in RIGHT_COLS else
                            "center" if col in CENTER_COLS else "left"),
                vertical="center",
                wrap_text=False,
            )

        ws.row_dimensions[excel_row].height = 18

    ws.freeze_panes = "A9"
    return make_output(wb)

# ── alias เพื่อ backward compatible กับโค้ดเดิม ──────────
def export_license_report(db, year=None, status=None):
    """alias → export_licensee_report"""
    return export_licensee_report(db, year=year, status=status)